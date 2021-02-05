from collections import Counter
from bs4 import BeautifulSoup
from urllib.parse import quote
import requests
import json
import time
import pandas as pd
import xlsxwriter
import openpyxl
import os
import traceback

dirname = os.path.dirname(__file__)
excel_filename = os.path.join(dirname, "betterbuys.xlsx")

# to create excel sheet
def create_sheet(data, sheet_name, sheet_no=1):
    try:
        print(sheet_no)
        df = pd.DataFrame(data)
        if sheet_no == 1:
            with pd.ExcelWriter(
                excel_filename, engine="openpyxl", mode="w"
            ) as writer:
                df.to_excel(writer, sheet_name=str(sheet_name), index=False)
        else:
            with pd.ExcelWriter(
                excel_filename, engine="openpyxl", mode="a"
            ) as writer:
                df.to_excel(writer, sheet_name=str(sheet_name), index=False)

    except Exception as e:
        print("Exception while get state commision", e)
        traceback.print_exc()

# create excel file
def create_new_excel():
    workbook = xlsxwriter.Workbook(excel_filename)
    workbook.close()


def soup_creator(response):
    return BeautifulSoup(response.text, "lxml")

# this function get return amazon product list based on user inputs
def get_data_from_amazon(input_dict, page_no=1, amazon_data_list=[]):
    try:

        headers = {
            "user-agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.102 Safari/537.36"
        }
        url = "https://www.amazon.in/s"
        sort_by = "relevancerank"
        if input_dict.get("sort_by") == "low_to_high":
            sort_by = "price-asc-rank"
        elif input_dict.get("sort_by") == "high_to_low":
            sort_by = "price-desc-rank"
        elif input_dict.get("sort_by") == "new arrivals":
            sort_by = "date-desc-rank"

        price_range_from = ""
        price_range_to = ""
        if input_dict.get("price_range_from", "") != "":
            price_range_from = str(
                int(input_dict.get("price_range_from")) * 100
            )

        if input_dict.get("price_range_to", "") != "":
            price_range_to = str(int(input_dict.get("price_range_to")) * 100)
        params = {
            "k": input_dict["product_name"],
            "s": sort_by,
            "rh": "n:976419031,p_36:"
            + price_range_from
            + "-"
            + price_range_to,
        }
        # print(params)
        if input_dict.get("page_no") is not None:
            params = {
                "k": input_dict["product_name"],
                "page": input_dict["page_no"],
                "s": sort_by,
                "rh": "n:976419031,p_36:"
                + price_range_from
                + "-"
                + price_range_to,
            }
        resp = requests.get(url, headers=headers, params=params)
        soup = soup_creator(resp)
        if soup.find("input", {"id": "captchacharacters"}):
            print("Captcha required")
            time.sleep(4)
            amazon_data_list = get_data_from_amazon(
                input_dict, amazon_data_list=amazon_data_list, page_no=page_no
            )
            return amazon_data_list
        with open("index.html", "w") as html:
            html.write(str(soup))
        products = soup.find_all(
            "div", {"data-component-type": "s-search-result"}
        )
        category = soup.find("ul", {"aria-labelledby": "n-title"})
        for product in products:
            data_dict = {}
            product_name = product.find(
                "h2",
                {
                    "class": "a-size-mini a-spacing-none a-color-base s-line-clamp-2"
                },
            )
            if product_name is not None:
                data_dict["product_name"] = (
                    product.find(
                        "h2",
                        {
                            "class": "a-size-mini a-spacing-none a-color-base s-line-clamp-2"
                        },
                    )
                    .text.strip()
                    .lower()
                )
            else:
                data_dict["product_name"] = (
                    product.find(
                        "h2",
                        {
                            "class": "a-size-mini a-spacing-none a-color-base s-line-clamp-4"
                        },
                    )
                    .text.strip()
                    .lower()
                )

            product_url = product.find(
                "h2",
                {
                    "class": "a-size-mini a-spacing-none a-color-base s-line-clamp-2"
                },
            )
            if product_url is not None:
                data_dict[
                    "product_url"
                ] = "https://www.amazon.in" + product.find(
                    "h2",
                    {
                        "class": "a-size-mini a-spacing-none a-color-base s-line-clamp-2"
                    },
                ).find(
                    "a"
                ).get(
                    "href"
                )
            else:
                data_dict[
                    "product_url"
                ] = "https://www.amazon.in" + product.find(
                    "h2",
                    {
                        "class": "a-size-mini a-spacing-none a-color-base s-line-clamp-4"
                    },
                ).find(
                    "a"
                ).get(
                    "href"
                )
            data_dict["source"] = "amazon"
            price = product.find("span", {"class": "a-offscreen"})
            if price is not None:
                data_dict["price"] = product.find(
                    "span", {"class": "a-offscreen"}
                ).text.strip()
            else:
                data_dict["price"] = "0"
            if category is None:
                data_dict["category"] = ""
            else:
                data_dict["category"] = category.find("li").text.strip()
            data_dict["product_image"] = product.find(
                "img", {"class": "s-image"}
            )["src"]
            data_dict["rating"] = (
                soup.find("div", {"class": "a-row a-size-small"})
                .text.split("out")[0]
                .strip()
            )
            if data_dict["price"] != "0":
                amazon_data_list.append(data_dict.copy())
            if len(amazon_data_list) >= input_dict["size"]:
                print(len(amazon_data_list), input_dict["size"])
                return amazon_data_list[0 : input_dict["size"]]
        else:
            page_no = page_no + 1
            input_dict["page_no"] = page_no
            print("request for next page")
            time.sleep(2)
            amazon_data_list = get_data_from_amazon(
                input_dict, page_no=page_no, amazon_data_list=amazon_data_list
            )
            return amazon_data_list[0 : input_dict["size"]]
    except Exception as e:
        print(amazon_data_list)
        print("Exception in get_data_from_amazon function", e)
        time.sleep(3)
        amazon_data_list = get_data_from_amazon(input_dict)
        return amazon_data_list[0 : input_dict["size"]]

# this function get return flipkart product list based on user inputs
def get_data_from_flipkart_json(input_dict, flipkart_data_list=[], page_no=1):
    try:
        headers = {
            "X-User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.102 Safari/537.36 FKUA/website/42/website/Desktop",
            "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.102 Safari/537.36",
            "Content-Type": "application/json",
        }
        sort_by = "relevance"
        if input_dict.get("sort_by") == "low_to_high":
            sort_by = "price_asc"
        elif input_dict.get("sort_by") == "high_to_low":
            sort_by = "price_desc"
        elif input_dict.get("sort_by") == "new_arrivals":
            sort_by = "recency_desc"

        parameters = (
            "/search?q="
            + input_dict["product_name"]
            + "&otracker=search&otracker1=search&marketplace=FLIPKART&as-show=off&as=off&sort="
            + sort_by
        )
        parameters += quote(
            "&p[]="
            + "facets.price_range.from="
            + str(input_dict.get("price_range_from"))
            + "&p[]="
            "facets.price_range.to=" + str(input_dict.get("price_range_to"))
        )
        serssion = requests.Session()
        data = (
            '{"pageUri":"'
            + parameters
            + '","pageContext":{"fetchSeoData":true,"paginatedFetch":false,"pageNumber":1},"requestContext":{"type":"BROWSE_PAGE","ssid":"","sqid":""}}'
        )
        resp = serssion.get("https://www.flipkart.com/")
        res = serssion.post(
            "https://1.rome.api.flipkart.com/api/4/page/fetch",
            headers=headers,
            cookies=resp.cookies.get_dict(),
            data=data,
        )

        for products in res.json()["RESPONSE"]["slots"][7:]:
            if products["widget"]["data"].get("products") is not None:
                for product in products["widget"]["data"]["products"]:
                    data_dict = {}
                    data_dict["product_name"] = product["productInfo"][
                        "value"
                    ]["titles"]["title"]
                    data_dict["price"] = "₹" + str(
                        int(
                            float(
                                product["productInfo"]["value"]["pricing"][
                                    "finalPrice"
                                ]["decimalValue"]
                            )
                        )
                    )
                    data_dict["product_url"] = (
                        "https://www.flipkart.com"
                        + product["productInfo"]["action"]["url"]
                    )
                    data_dict["source"] = "flipkart"
                    data_dict["product_image"] = (
                        product["productInfo"]["value"]["media"]["images"][0][
                            "url"
                        ]
                        .replace("{@width}", "612")
                        .replace("{@width}", "612")
                        .replace("{@height}", "612")
                        .replace("q={@quality}", "")
                    )
                    data_dict["rating"] = product["productInfo"]["value"][
                        "rating"
                    ]["average"]
                    data_dict["category"] = product["productInfo"]["value"][
                        "analyticsData"
                    ]["category"]
                    flipkart_data_list.append(data_dict.copy())
                    if len(flipkart_data_list) >= input_dict["size"]:
                        return flipkart_data_list
        else:
            page_no += 1
            flipkart_data_list = get_data_from_flipkart_json(
                input_dict,
                flipkart_data_list=flipkart_data_list,
                page_no=page_no,
            )
            return flipkart_data_list
    except Exception as e:
        print("Exception in get_data_from_flipkart_json", e)
        time.sleep(4)
        print(res.json())
        flipkart_data_list = get_data_from_flipkart_json(
            input_dict, flipkart_data_list=flipkart_data_list, page_no=page_no
        )
        return flipkart_data_list

# this function process data and sort to price wise. 
def get_product_data_sort_wise(dataList, input_dict):
    dataList = dataList[: input_dict["size"]]
    key_counts = Counter(d["product_name"] for d in dataList)
    uniqueValues = []
    duplicateValues = []
    for d in dataList:
        if key_counts[d["product_name"]] == 1:
            uniqueValues.append(d)
        else:
            if not any(
                d["product_name"] == k["product_name"] for k in duplicateValues
            ):
                duplicateValues.append(d)

    sort_list = sorted(
        duplicateValues,
        key=lambda i: int(
            i["price"].split(".")[0].replace("₹", "").replace(",", "")
        ),
    )
    sortuniqueValues = sorted(
        uniqueValues,
        key=lambda i: int(
            i["price"].split(".")[0].replace("₹", "").replace(",", "")
        ),
    )

    sort_list.extend(sortuniqueValues)
    create_new_excel()
    create_sheet(dataList, "sheet")
    create_sheet(sort_list, "sheet2", sheet_no=2)

# validate user input information
def validate_params(input_dict):
    if (
        input_dict.get("size") == ""
        or input_dict.get("size") is None
        or int(input_dict.get("size", 0)) <= 10
    ):
        input_dict["size"] = 10
    elif int(input_dict.get("size")) > 50:
        input_dict["size"] = 50
    return input_dict

# this function return split divide data for how many data get from amazon and flipkart
def split_data_size_wise(input_dict):
    size = int(input_dict["size"])
    flipkart_size = int(size / 2)
    amazon_size = size - flipkart_size
    return flipkart_size, amazon_size

# here start the scraping process
def start_parsing(input_dict):
    input_dict = validate_params(input_dict)
    flipkart_size, amazon_size = split_data_size_wise(input_dict)
    input_dict["size"] = amazon_size
    amazon_data = get_data_from_amazon(input_dict)
    flipkart_data = get_data_from_flipkart_json(input_dict)
    flipkart_data.extend(amazon_data)
    input_dict["size"] = amazon_size + flipkart_size
    get_product_data_sort_wise(flipkart_data, input_dict)

